"""Tests for helpers for Excel files (using openpyxl)."""

import os
import os.path as pth
import tempfile
import uuid

from contextlib import contextmanager

from nose.tools import eq_
from openpyxl import Workbook

from datasimple.xl import (
    ValueMapper,
    XlsxImporter,
    ws_scan,
    ws_scan_raw,
    ws_sheet_names,
    ics_report_params
)

CONFIG = """
[WORKBOOK]
Col1 = IMNormal
Col2 = IMComma
Col3 = IMCurrency
IntCol = IMInt
FloatCol = IMFloat
WCCol* = IMInt

[Sheet1]
Col1 = IMCurrency
Col2 = IMCurrency

[AllNormal]
Qty = IMNormal
ExtPrice = IMNormal
"""


def eqf_(f1, f2):
    """Helper for comparing floats."""
    if abs(f1 - f2) > 0.000001:
        raise ValueError('{} != {}'.format(f1, f2))


def eqfm_(a, b):
    """Helper for comparing floats AND style names."""
    n1, v1 = a
    n2, v2 = b
    if type(v1) is not float:
        return eq_(a, b)
    eqf_(v1, v2)
    eq_(n1, n2)


def mapper_default_test():
    map_src = ValueMapper()
    m = map_src.create_mapper('NopeUseDefault')

    eq_(('IMComma', 'A'), m('Qty', 'A'))
    eq_(('IMComma', 0), m('Qty', '0'))
    eq_(('IMComma', 0), m('Qty', 0))
    eq_(('IMComma', 12), m('Qty', '12'))
    eq_(('IMComma', 12), m('Qty', 12))
    eq_(('IMComma', 12), m('Qty', '12.0'))
    eq_(('IMComma', 12), m('Qty', 12.0))

    eqfm_(('IMCurrency', 'A'),  m('ExtPrice', 'A'))
    eqfm_(('IMCurrency', 0.0), m('ExtPrice', '0'))
    eqfm_(('IMCurrency', 0.0), m('ExtPrice', 0))
    eqfm_(('IMCurrency', 12.0), m('ExtPrice', '12'))
    eqfm_(('IMCurrency', 12.0), m('ExtPrice', 12))
    eqfm_(('IMCurrency', 12.0), m('ExtPrice', '12.0'))
    eqfm_(('IMCurrency', 12.0), m('ExtPrice', 12.0))

    eq_(('IMNormal', 'A'),  m('IDontKnow', 'A'))
    eq_(('IMNormal', '12'), m('IDontKnow', '12'))
    eq_(('IMComma', 12), m('IDontKnow', 12))
    eq_(('IMNormal', '12.0'), m('IDontKnow', '12.0'))
    eqfm_(('IMCurrency', 12.0), m('IDontKnow', 12.0))


@contextmanager
def temp_xlsx_name():
    """Python temp files assume that we will use the file handle, but we need a name
    for openpyxl. So we add our own uniqifier and get the temp name."""
    suffix = uuid.uuid4().hex + '.xlsx'
    with tempfile.NamedTemporaryFile(mode='w', suffix=suffix, delete=True) as fp:
        fp.flush()
        tmpname = fp.name
        assert tmpname

    try:
        # If we made it here, then we were able to create and delete the temp file
        yield tmpname
    finally:
        try:
            os.remove(tmpname)
        except Exception:
            print('Error deleting named temp file ' + tmpname)


def ics_parms_test():
    PARMS = [
        'Parm1: First',
        'ParmMult: A',
        'ParmMult: B',
        'ParmMult: C',
        'Parm42: Last',
    ]
    wb = Workbook()
    sheet = wb.create_sheet('CustomCover')
    sheet.cell(row=1, column=1, value='Parms')
    for idx, val in enumerate(PARMS):
        sheet.cell(row=2+idx, column=1, value=val)

    with temp_xlsx_name() as tmpname:
        wb.save(tmpname)
        read = ics_report_params(tmpname, sheet_name='CustomCover')
        eq_('MISSING', read.get('ParmX', 'MISSING'))
        eq_('First', read['Parm1'])
        eq_('Last', read['Parm42'])
        eq_(['A', 'B', 'C'], read['ParmMult'])


def mapper_config_test():
    with tempfile.NamedTemporaryFile(mode='w+', suffix='.xlsx', delete=False) as fp:
        fp.write(CONFIG)
        fp.flush()
        map_src = ValueMapper(fp.name)

    # Override default cols
    m = map_src.create_mapper('AllNormal')
    eq_(('IMNormal', 'A'), m('Qty', 'A'))
    eq_(('IMNormal', '1'), m('Qty', '1'))
    eq_(('IMNormal', 1),   m('Qty', 1))

    eq_(('IMNormal', 'A'), m('ExtPrice', 'A'))
    eq_(('IMNormal', '1'), m('ExtPrice', '1'))
    eq_(('IMNormal', 1),   m('ExtPrice', 1))

    eq_(('IMNormal', 'A'), m('ColNotThere', 'A'))
    eq_(('IMNormal', '1'), m('ColNotThere', '1'))
    eq_(('IMComma', 1),   m('ColNotThere', 1))

    # Use default section
    m = map_src.create_mapper('NopeUseDefault')
    eq_(('IMNormal', 'A'), m('Col1', 'A'))

    eq_(('IMComma', 'A'), m('Col2', 'A'))
    eq_(('IMComma', 12), m('Col2', '12'))

    eqfm_(('IMCurrency', 'A'), m('Col3', 'A'))
    eqfm_(('IMCurrency', 21.0), m('Col3', '21'))

    eq_(('IMNormal', 'A'), m('ColNotThere', 'A'))
    eq_(('IMNormal', '1'), m('ColNotThere', '1'))
    eq_(('IMComma', 1), m('ColNotThere', 1))

    eq_(('IMComma', 1), m('Qty', '1'))
    eqfm_(('IMCurrency', 1.0),  m('ExtPrice', '1'))

    # Use default + sheet sections
    m = map_src.create_mapper('Sheet1')
    eqfm_(('IMCurrency', 'A'),  m('Col1', 'A'))
    eqfm_(('IMCurrency', 42.2), m('Col1', '42.2'))
    eqfm_(('IMCurrency', 42.2), m('Col1', 42.2))
    eqfm_(('IMCurrency', 42.0), m('Col1', 42))

    eqfm_(('IMCurrency', 'A'),  m('Col2', 'A'))
    eqfm_(('IMCurrency', 42.2), m('Col2', '42.2'))
    eqfm_(('IMCurrency', 42.2), m('Col2', 42.2))
    eqfm_(('IMCurrency', 42.0), m('Col2', 42))

    eqfm_(('IMCurrency', 'A'),  m('Col3', 'A'))
    eqfm_(('IMCurrency', 42.2), m('Col3', '42.2'))
    eqfm_(('IMCurrency', 42.2), m('Col3', 42.2))
    eqfm_(('IMCurrency', 42.0), m('Col3', 42))

    eq_(('IMNormal', 'A'), m('ColNotThere', 'A'))
    eq_(('IMNormal', '1'), m('ColNotThere', '1'))

    eq_(('IMComma', 1), m('Qty', '1'))
    eqfm_(('IMCurrency', 1.1),  m('ExtPrice', '1.1'))

    # Make sure we test the other available formats
    eq_(('IMInt', 'A'), m('IntCol', 'A'))
    eq_(('IMInt', 0),   m('IntCol', '0'))
    eq_(('IMInt', 0),   m('IntCol', 0))
    eq_(('IMInt', 12),  m('IntCol', '12'))
    eq_(('IMInt', 12),  m('IntCol', 12))
    eq_(('IMInt', 12),  m('IntCol', '12.0'))
    eq_(('IMInt', 12),  m('IntCol', 12.0))

    eqfm_(('IMFloat', 'A'),  m('FloatCol', 'A'))
    eqfm_(('IMFloat', 0.0),  m('FloatCol', '0'))
    eqfm_(('IMFloat', 0.0),  m('FloatCol', 0))
    eqfm_(('IMFloat', 12.0), m('FloatCol', '12'))
    eqfm_(('IMFloat', 12.0), m('FloatCol', 12))
    eqfm_(('IMFloat', 12.0), m('FloatCol', '12.0'))
    eqfm_(('IMFloat', 12.0), m('FloatCol', 12.0))

    # Wildcards should match exactly AND with globbinb
    eq_(('IMInt', 12),  m('WCCol*', '12'))
    eq_(('IMInt', 12),  m('WCCol-SomeExtraStuff', '12'))


def importer_tests():
    class TestImporter(XlsxImporter):
        def add_args(self, argparser):
            argparser.add_argument('-x', '--testingarg', required=True, help='Test parm')

        def validate_args(self, args):
            assert args.testingarg == 'Saw Arg', 'Arg not passed thru'

        def customize_val_mapper(self, value_mapper):
            value_mapper.default_mapping['SomeInt'] = 'IMComma'
            value_mapper.default_mapping['SomeFloat'] = 'IMCurrency'

        def get_data(self, args):
            def rows():
                yield ['r1', '42', '42.42']
                yield ['r2', '43', '43.43']
                yield ['r3', '-44', '-44.44']
            return ['Name', 'SomeInt', 'SomeFloat'], rows()

    with tempfile.TemporaryDirectory() as folder:
        fn = pth.join(folder, 'vanilla_file.xlsx')
        args = [
            '-b', fn,
            '-s', 'TestSheet',
            '-f', 'A2',
            '-x', 'Saw Arg'
        ]

        assert not pth.isfile(fn), 'Previous file found? Test can not run'

        def check_sheet():
            TestImporter().main(cmdline_args=args)

            assert pth.isfile(fn), 'File save failed'

            eq_(['TestSheet'], list(ws_sheet_names(fn)))

            rows = list(ws_scan(fn, 'TestSheet'))
            eq_(3, len(rows))

            eq_('r1', rows[0]['Name'])
            eq_('r2', rows[1]['Name'])
            eq_('r3', rows[2]['Name'])

            eq_(42, int(rows[0]['SomeInt']))
            eq_(43, int(rows[1]['SomeInt']))
            eq_(-44, int(rows[2]['SomeInt']))

            eqf_(42.42,  float(rows[0]['SomeFloat']))
            eqf_(43.43,  float(rows[1]['SomeFloat']))
            eqf_(-44.44, float(rows[2]['SomeFloat']))

        # Check twice - once for create and once for rewrite
        check_sheet()
        check_sheet()


def importer_tranpose_tests():
    class TestImporter(XlsxImporter):
        def add_args(self, argparser):
            pass

        def validate_args(self, args):
            pass

        def customize_val_mapper(self, value_mapper):
            pass

        def get_data(self, args):
            def rows():
                yield ['1,1', '1,2']
                yield ['2,1', '2,2']
            return ['c1', 'c2'], rows()

    with tempfile.TemporaryDirectory() as folder:
        fn = pth.join(folder, 'transpose_file.xlsx')
        args = [
            '-b', fn,
            '-s', 'TransposeSheet',
            '--transpose'
        ]

        assert not pth.isfile(fn), 'Previous file found? Test can not run'

        def check_sheet():
            TestImporter().main(cmdline_args=args)

            assert pth.isfile(fn), 'File save failed'

            rows = list(ws_scan_raw(fn, 'TransposeSheet'))
            eq_(2, len(rows))
            eq_(['c1', '1,1', '2,1'], rows[0])
            eq_(['c2', '1,2', '2,2'], rows[1])

        # Check twice - once for create and once for rewrite
        check_sheet()
        check_sheet()
