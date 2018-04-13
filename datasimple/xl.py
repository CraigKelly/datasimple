"""Provide helpers for Excel files (using openpyxl)."""

# TODO: Check and see if openpyxl is upto version 2.5 yet - that should include the
#       file handle leak that is cluttering the test output with Resource Warnings

import glob
import os.path as pth
import sys

from argparse import ArgumentParser
from collections import ChainMap
from contextlib import closing
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill

from .core import norm_ws, kv, read_config
from .cli import log


# Luckily the functionality we want to already out there
globmatch = glob.fnmatch.fnmatch


def _val(cell):
    v = cell.value
    if v is None:
        return ''

    t = type(v)
    if t is str:
        if v and v[0] == "'" and v.find("'", 1) < 0:
            v = v[1:]  # starts with ' and doesn't have a match: old excel "force string" method
        return norm_ws(v)
    elif t is int:
        return v
    elif t is float:
        return v
    elif t is datetime:
        return v
    else:
        return norm_ws(repr(v))


# This is generally just for us, but some might find it useful
def load_ro_workbook(xlsx_file):
    """Read a workbook opened read-only for fast access."""
    return load_workbook(filename=xlsx_file, read_only=True, data_only=True)


def ws_sheet_names(xlsx_file, log_on_open=True):
    """Return the sheet names in the XLSX file."""
    if log_on_open:
        log('SCAN: [!c]{:s}[!/c]', xlsx_file)
    with closing(load_ro_workbook(xlsx_file)) as wb:
        return [str(i) for i in wb.sheetnames]


def ws_scan_raw(xlsx_file, sheet_name, log_on_open=True):
    """Iterator for every row in sheet_name in xlsx_file."""
    if log_on_open:
        log('OPEN: [!c]{:s} => {:s}[!/c]', xlsx_file, sheet_name)
    with closing(load_ro_workbook(xlsx_file)) as wb:
        ws = wb[sheet_name]
        for row in ws:
            yield [_val(cell) for cell in row]


def ws_scan(xlsx_file, sheet_name, log_on_open=True):
    """Iterator for every row in sheet_name in xlsx_file, returned as dict."""
    headers = None
    for vals in ws_scan_raw(xlsx_file, sheet_name, log_on_open=log_on_open):
        if not headers:
            headers = vals
            continue
        yield dict((k, v) for k, v in zip(headers, vals) if k)


def add_named_style(
    wb,
    name,
    number_format=None,
    font_name='Calibri',
    font_size=10,
    font_color='FF000000',
    fill=None
):
    """Add a named style to the open workbook.

    Already named styles are ignored. See add_default_styles for example usage.
    """
    if name in wb.named_styles:
        return
    sty = NamedStyle(name=name)
    sty.font = Font(name=font_name, size=font_size, color=font_color)
    if fill:
        sty.fill = fill
    if number_format:
        sty.number_format = number_format
    wb.add_named_style(sty)


def add_default_styles(wb):
    """Add default styles we use in some tools"""
    add_named_style(wb, 'IMHeader', font_color='FFFFFFFF', fill=PatternFill('solid', fgColor='FF4F81BD'))
    add_named_style(wb, 'IMNormal')
    add_named_style(wb, 'IMInt', number_format='0')
    add_named_style(wb, 'IMFloat', number_format='0.0000')
    add_named_style(wb, 'IMComma', number_format='#,##0')
    add_named_style(wb, 'IMPercent', number_format='0.000%')
    add_named_style(wb, 'IMCurrency', number_format='_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)')
    add_named_style(wb, 'IMDate', number_format='m/d/yy')


def ics_report_params(xlsx_file, sheet_name='Cover'):
    """Return a dictionary of report parameters for an ICS batch system report.
    If a parameter appears more than once, the values are returned as a list"""
    parms = dict()
    for vals in ws_scan(xlsx_file, sheet_name):
        val = vals['Parms']
        k, v = kv(val)
        if not k:
            continue  # invalid line

        if k in parms:
            # We have a multi-value parameter
            if type(parms[k]) is not list:
                parms[k] = [parms[k]]
            parms[k].append(v)
        else:
            # Just a normal, single val parameter
            parms[k] = v

    return parms


def _int(v):
    """Convert to int for excel, but default to original value."""
    try:
        if v is None or v == '':
            return ''
        if type(v) is str:
            # we handle strings like '2,345.00'
            return int(float(v.replace(',', '')))
        return int(v)
    except ValueError:
        return v


def _acct(v):
    """Convert to accounting (float) for excel, but default to original value."""
    try:
        if v is None or v == '':
            return ''
        return float(v)
    except ValueError:
        return v


class ValueMapper(object):
    """Support column-name-based value casting and cell formatting.

    After creation, call create_mapper with a sheet name to create
    a mapper function. That function takes a column name and a
    value then returns (StyleName, value)."""
    def __init__(self, config_file=None):
        self.default_mapping = {
            'Qty': 'IMComma',
            'ExtPrice': 'IMCurrency',
        }

        self.config_file = config_file
        if self.config_file:
            with open(self.config_file) as inp:
                cfg_text = str(inp.read()).strip()
            self.config = read_config(cfg_text)
        else:
            self.config = dict()

        self.type_map = {
            int: 'IMComma',
            float: 'IMCurrency',
        }

        self.convert_map = {
            'IMComma': _int,
            'IMInt': _int,
            'IMCurrency': _acct,
            'IMFloat': _acct,
            'IMPercent': _acct,
        }

    def create_mapper(self, sheet_name, log_choices=False):
        col_map = ChainMap(
            self.config.get(sheet_name, {}),
            self.config.get('WORKBOOK', {}),
            self.default_mapping
        )
        if log_choices:
            log('Col Map for {}: {}', sheet_name, col_map)
            log('CONFIG: {}', self.config)
        type_map = dict(self.type_map)
        convert_map = dict(self.convert_map)

        # We allow wildcards in the column names now
        wildcards = list((k.lower(), v) for k, v in col_map.items() if '*' in k)

        def m(col_name, val):
            # See if there's a col name mapping
            style_name = col_map.get(col_name, '')
            src = 'col_map'

            if not style_name:
                # See if there's a wildcard match in the col mappings
                lcn = col_name.lower()
                for patt, sty in wildcards:
                    if globmatch(lcn, patt):
                        style_name, src = sty, 'col_map:wildcard'
                        break

            if not style_name:
                # Punt based on type
                style_name = type_map.get(type(val), '')
                src = 'type_map'

            if not style_name:
                # Still stumped: just default
                style_name = 'IMNormal'
                src = 'default_style'

            # We perform some conversions automatically
            conv = convert_map.get(style_name, None)
            if conv:
                val = conv(val)

            if log_choices:
                log('{}, {} => {} via {}', col_name, val, style_name, src)

            # Finally done
            return style_name, val

        return m


class XlsxImporter(object):
    """Base class used to create xlsx import scripts.update_metadata

    This is for our csv2xlsx and sql2xlsx tools - and hopefully future stuff as
    well. It should also be handy for custom spreadsheet creation.
    """

    def __init__(self):
        """Construction."""
        pass

    def add_args(self, argparser):
        """Add any arguments parser arguments necessary."""
        raise NotImplementedError

    def validate_args(self, args):
        """Validate values for custom arguments given in add_args."""
        raise NotImplementedError

    def customize_val_mapper(self, value_mapper):
        """Optional way to customize ValueMapper instance."""
        pass

    def get_data(self, args):
        """Return tuple of (cols, rows).

        Cols is an iterable of column names.
        Rows is an iterable of rows where each row is an iterable of values.

        Each row should correspond to the column list each.
        """
        raise NotImplementedError

    def before_save(self, args, wb, sheet):
        """Optional last chance at the sheet before save."""
        pass

    def main(self, cmdline_args=None):
        """Our main contribution: this is the logic that we provide."""
        # Our default arguments
        parser = ArgumentParser()
        parser.add_argument('-b', '--book',      type=str, required=True, help='Name of workbook to create/update')
        parser.add_argument('-s', '--sheetname', type=str, required=True, help='Name of worksheet to create')
        parser.add_argument('-m', '--mapper',    type=str, default='',    help='Mapper config file to use')
        parser.add_argument('-f', '--freeze',    type=str, default='',    help='Optional cell at which to perform a Freeze Panes (e.g. use A2 to freeze top row)')
        parser.add_argument('-t', '--transpose', action='store_true', default=False, help='If set, transpose result sheet')

        # Get any necessary arguments, parse everything, and then perform
        # init validation
        self.add_args(parser)
        if cmdline_args is None:
            cmdline_args = list(sys.argv[1:])
        args = parser.parse_args(args=cmdline_args)
        self.validate_args(args)

        # Create col/value mapper
        if args.mapper:
            assert pth.isfile(args.mapper)
        mapper_src = ValueMapper(args.mapper)
        self.customize_val_mapper(mapper_src)
        mapper = mapper_src.create_mapper(args.sheetname)

        # Create or open workbook and get our worksheet ready
        if pth.isfile(args.book):
            log('Opening [!y]{:s}[!/y]', args.book)
            wb = load_workbook(args.book)
        else:
            log('Creating [!y]{:s}[!/y]', args.book)
            wb = Workbook()

        wb.guess_types = False

        if args.sheetname in wb.sheetnames:
            log('Removing previous sheet [!r]{:s}[!/r]', args.sheetname)
            del wb[args.sheetname]

        log('Creating worksheet [!y]{:s}[!/y]', args.sheetname)
        sheet = wb.create_sheet(args.sheetname)

        # openpyxl creates a default worksheet named 'Sheet': remove it unless that's
        # what we just created...
        if args.sheetname != 'Sheet' and 'Sheet' in wb.sheetnames:
            log('Removing DEFAULT SHEET named [!r]Sheet[!/r]')
            del wb['Sheet']

        # Add our standard styles
        add_default_styles(wb)

        # Now we need the data that we'll be writing
        col_names, rows = self.get_data(args)

        # simplify cell writing, and handle transpoition
        def _write_cell(r, c, v, sty):
            if args.transpose:
                r, c = c, r
            sheet.cell(row=r, column=c, value=v).style = sty

        # Create header row
        col_names = list(col_names)  # Go ahead and freeze column names
        for idx, col in enumerate(col_names):
            _write_cell(1, idx+1, col, 'IMHeader')

        # Now create all rows
        count = 0
        for row in rows:
            for idx, val in enumerate(row):
                col = col_names[idx]
                style, val = mapper(col, val)
                _write_cell(count+2, idx+1, val, style)

            count += 1
            if count == 1:
                log('[!g]First Record Written![!/g]')
            elif count % 5000 == 0:
                log('Records: [!g]{:,d}[!/g]', count)

        # Finalize sheet - we autofit cols, freeze if necessary, and save
        for col in sheet.columns:
            column = col[0].column   # need column name for below
            max_length = 6
            for cell in col:
                v = cell.value
                if v:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 3.2) * 0.88  # calc is totally arbitrary
            sheet.column_dimensions[column].width = adjusted_width

        # Freeze if requests
        if args.freeze:
            log('Freezing sheet at [!c]{}[!/c]', args.freeze)
            sheet.freeze_panes = args.freeze

        # Give our implementor one last shot
        self.before_save(args, wb, sheet)

        log('[!c]Saving[!/c]')
        wb.save(args.book)
        wb.close()

        log('[!br][!w]DONE[!/w][!/br] -> Rows: [!g]{:,d}[!/g]', count)
